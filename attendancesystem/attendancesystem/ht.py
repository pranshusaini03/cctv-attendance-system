import cv2
import numpy as np
from tensorflow.keras.models import load_model
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

model = load_model(r"face_recognition_model.h5")
IMG_SIZE = (128, 128)  # Must match the training image size
images_path = r"C:\Users\Pranshu Saini\Desktop\images"
class_names = {index: name for index, name in enumerate(sorted(os.listdir(images_path)))}

model_file = r"res10_300x300_ssd_iter_140000.caffemodel"
config_file = r"deploy.prototxt"
dnn_net = cv2.dnn.readNetFromCaffe(config_file, model_file)

excel_file = r"attendance.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Confidence", "Timestamp"])
    wb.save(excel_file)

# Button position (bottom right of the frame)
button_x1, button_y1, button_x2, button_y2 = 450, 400, 600, 450
recognized_name = "Unknown"
probability_value = 0.0

def preprocess_image(img):
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img_resized = cv2.resize(img_rgb, IMG_SIZE)
    img_normalized = img_resized / 255.0
    img_reshaped = np.reshape(img_normalized, (1, IMG_SIZE[0], IMG_SIZE[1], 3))
    return img_reshaped

def save_attendance():
    """Save recognized name to Excel."""
    global recognized_name, probability_value
    if recognized_name != "Unknown":
        wb = load_workbook(excel_file)
        ws = wb.active
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([recognized_name, round(probability_value * 100, 2), timestamp])
        wb.save(excel_file)
        print(f"Attendance saved for: {recognized_name}")

def mouse_callback(event, x, y, flags, param):
    """Detect button click."""
    if event == cv2.EVENT_LBUTTONDOWN:
        if button_x1 < x < button_x2 and button_y1 < y < button_y2:
            save_attendance()

def take_attendanceht(device):
    cap = cv2.VideoCapture(0)
    cap.set(3, 640)
    cap.set(4, 480)
    cv2.namedWindow("Face Recognition")
    cv2.setMouseCallback("Face Recognition", mouse_callback)
    if(device==1):
        address = "http://172.20.10.7:8080/video"
        cap.open(address)
    
    global recognized_name, probability_value

    while True:
        success, frame = cap.read()
        if not success:
            break

        blob = cv2.dnn.blobFromImage(frame, 1.0, (300, 300), (104.0, 177.0, 123.0))
        dnn_net.setInput(blob)
        detections = dnn_net.forward()

        for i in range(detections.shape[2]):
            confidence = detections[0, 0, i, 2]
            if confidence > 0.5:
                box = detections[0, 0, i, 3:7] * np.array([frame.shape[1], frame.shape[0], frame.shape[1], frame.shape[0]])
                (x, y, x2, y2) = box.astype("int")
                x, y, x2, y2 = max(0, x), max(0, y), min(frame.shape[1], x2), min(frame.shape[0], y2)
                cropped_face = frame[y:y2, x:x2]
                processed_face = preprocess_image(cropped_face)

                prediction = model.predict(processed_face)
                class_index = np.argmax(prediction)
                probability_value = prediction[0][class_index]

                if probability_value > 0.5:
                    recognized_name = class_names[class_index]
                    label = f"{recognized_name}: {round(probability_value * 100, 2)}%"
                else:
                    recognized_name = "Unknown"
                    label = "Unknown"

                cv2.rectangle(frame, (x, y), (x2, y2), (0, 255, 0), 2)
                cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        # Draw button
        cv2.rectangle(frame, (button_x1, button_y1), (button_x2, button_y2), (50, 50, 255), -1)
        cv2.putText(frame, "Attendance", (button_x1 + 5, button_y1 + 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

        cv2.imshow("Face Recognition", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

