import cv2
import numpy as np
from tensorflow.keras.models import load_model
from mtcnn import MTCNN
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# Load the trained model
model = load_model(r"face_recognition_model.h5")

# Parameters
IMG_SIZE = (128, 128)  # Must match the training image size
images_path = r"C:\Users\Pranshu Saini\Desktop\images"
class_names = {index: name for index, name in enumerate(sorted(os.listdir(images_path)))}

# Initialize MTCNN
mtcnn_detector = MTCNN()

# Attendance file path
excel_file = r"attendance.xlsx"

# Create an attendance file if it doesn't exist
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Confidence", "Timestamp"])
    wb.save(excel_file)

# Function to preprocess the image for the model
def preprocess_image(img):
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # Convert BGR to RGB
    img_resized = cv2.resize(img_rgb, IMG_SIZE)    # Resize to model input size
    img_normalized = img_resized / 255.0           # Normalize pixel values
    img_reshaped = np.reshape(img_normalized, (1, IMG_SIZE[0], IMG_SIZE[1], 3))  # Add batch dimension
    return img_reshaped

# Global variables for storing detected face details
recognized_name = None
confidence_value = None

# Mouse callback function for detecting button clicks
def click_event(event, x, y, flags, param):
    global recognized_name, confidence_value
    if event == cv2.EVENT_LBUTTONDOWN:

        if 400 <= x <= 450 and 450 <= y <= 600:  # Button region
            if recognized_name and confidence_value:
                wb = load_workbook(excel_file)
                ws = wb.active
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([recognized_name, confidence_value, timestamp])
                wb.save(excel_file)
                print(f"Attendance saved for {recognized_name}")

# Function to start face recognition and capture attendance
def take_attendancemtcnn(device):
    global recognized_name, confidence_value

    # Start webcam
    cap = cv2.VideoCapture(0)
    cap.set(3, 640)  # Set width
    cap.set(4, 480)  # Set height
    cv2.namedWindow("Real-Time Face Recognition")
    cv2.setMouseCallback("Real-Time Face Recognition", click_event)
    if(device==1):
        address = "http://172.20.10.7:8080/video"
        cap.open(address)
    
    while True:
        success, frame = cap.read()
        if not success:
            break

        # Convert frame to RGB for MTCNN
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # Using MTCNN for Face Detection
        faces = mtcnn_detector.detect_faces(rgb_frame)

        for result in faces:
            x, y, width, height = result['box']
            confidence = result['confidence']

            if confidence > 0.5:  # Confidence threshold
                cropped_face = frame[y:y+height, x:x+width]  # Crop the detected face
                processed_face = preprocess_image(cropped_face)  # Preprocess for the model

                # Predict using the model
                prediction = model.predict(processed_face)
                class_index = np.argmax(prediction)
                probability_value = round(prediction[0][class_index] * 100, 2)

                # Label face if confidence is high
                if probability_value > 50:
                    recognized_name = class_names[class_index]
                    confidence_value = probability_value
                    label = f"{recognized_name}: {probability_value}%"
                else:
                    label = "Unknown"
                    recognized_name = None
                    confidence_value = None

                # Draw rectangle and label around the face
                cv2.rectangle(frame, (x, y), (x + width, y + height), (0, 255, 0), 2)
                cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        # Draw "Save Attendance" button
        cv2.rectangle(frame, (50, 400), (200, 450), (0, 255, 255), -1)  # Button background
        cv2.putText(frame, "Attendance", (60, 430), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)

        cv2.imshow("Real-Time Face Recognition", frame)

        # Exit loop when 'q' key is pressed
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the webcam and close all OpenCV windows
    cap.release()
    cv2.destroyAllWindows()

    return "Attendance Taken Successfully!"

# Run the function
