from django.shortcuts import render
from django.contrib import messages
from .dnn import take_attendance  # Import the function from s.py or z.py
from .mtcn import take_attendancemtcnn
from .ht import take_attendanceht
from django.http import HttpResponse
from openpyxl import load_workbook
from .model import train_model
from .y import newstudent
import threading
from django.contrib import messages
from django.shortcuts import render
def main(request):
    return render(request, 'main.html')

def attendance(request):
    return render(request, 'attendance.html')

def startattendancednn(request,device):
    result=take_attendance(device)
    messages.success(request, "Task completed successfully!")
    return render(request, 'main.html')
def register(request):
    return render(request, 'register.html')

def view_attendance(request):
    excel_file = r"C:\Users\Pranshu Saini\Desktop\attendancesystem\attendance.xlsx"
    try:
            wb = load_workbook(excel_file)
            ws = wb.active  # Select the active worksheet

            # Read the Excel data into a list of dictionaries
            attendance_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
                attendance_data.append({
                    "name": row[0],
                    "confidence": row[1],
                    "timestamp": row[2]
                })

    except Exception as e:
            return HttpResponse(f"Error reading the Excel file: {str(e)}")

    # Pass the data to the template
    return render(request, 'view.html', {"attendance_data": attendance_data})
def new_student(request):
    if request.method == "POST":
        # Get the student's name from the POST request
        student_name = request.POST.get("student_name")
        
        # Ensure the name is not empty
        if student_name:
            # Call the function to start attendance, passing the student's name if required
            result = newstudent(student_name)  # Assuming newstudent() can handle a name
            # Render the result or a success message
            return render(request, 'model.html', {'message': f"Attendance started for {student_name}"})
        else:
            # Handle the case where no name is provided
            return render(request, 'model.html', {'error': "Please enter a valid name."})
    else:
        # For GET requests, render the form page
        return render(request, 'model.html')

def model(request):
    image_dir = r"C:\Users\Pranshu Saini\Desktop\images"  # Update this with the actual image directory path
    try:
        train_model(image_dir)
        messages.success(request, "Model trained successfully!")
    except Exception as e:
        messages.error(request, f"Error during training: {str(e)}")
    return render(request,'main.html')
def startattendancemtcnn(request,device):
    result=take_attendancemtcnn(device)
    messages.success(request, "Task completed successfully!")
    return render(request, 'main.html')
def startattendanceht(request,device):
    result=take_attendanceht(device)
    messages.success(request, "Task completed successfully!")
    return render(request, 'main.html')